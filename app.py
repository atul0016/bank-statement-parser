import os
import uuid
import traceback
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from parsers.detector import detect_and_parse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import json

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload

# Use /tmp on Render (read-only filesystem outside /tmp)
if os.environ.get('RENDER'):
    UPLOAD_FOLDER = '/tmp/uploads'
    OUTPUT_FOLDER = '/tmp/outputs'
else:
    UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
    OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), 'outputs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/parse', methods=['POST'])
def parse_pdf():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    password = request.form.get('password', '')

    # Save uploaded file
    ext = os.path.splitext(file.filename)[1]
    unique_name = str(uuid.uuid4()) + ext
    filepath = os.path.join(UPLOAD_FOLDER, unique_name)
    file.save(filepath)

    try:
        result = detect_and_parse(filepath, password=password if password else None)

        if result is None or len(result['transactions']) == 0:
            return jsonify({'error': 'Could not parse any transactions from this PDF. The bank format may not be supported or the PDF may be password-protected.'}), 400

        return jsonify({
            'bank_type': result['bank_type'],
            'account_info': result.get('account_info', ''),
            'transactions': result['transactions'],
            'filename': file.filename
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error parsing PDF: {str(e)}'}), 500
    finally:
        # Clean up uploaded file
        try:
            os.remove(filepath)
        except OSError:
            pass


@app.route('/export', methods=['POST'])
def export_excel():
    data = request.get_json()
    if not data or 'transactions' not in data:
        return jsonify({'error': 'No data to export'}), 400

    transactions = data['transactions']
    filename = data.get('filename', 'bank_statement') 
    bank_type = data.get('bank_type', 'Unknown')

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bank Statement'

    # Styling
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'Bank Statement - {bank_type}'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Source file info
    ws.merge_cells('A2:G2')
    ws['A2'].value = f'Source: {filename}'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['S.No.', 'Date', 'Description', 'Debit', 'Credit', 'Running Balance', 'Ledger']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for i, txn in enumerate(transactions):
        row = i + 5
        ws.cell(row=row, column=1, value=txn.get('serial', i + 1)).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=2, value=txn.get('date', '')).border = thin_border
        
        desc_cell = ws.cell(row=row, column=3, value=txn.get('description', ''))
        desc_cell.border = thin_border
        desc_cell.alignment = cell_alignment

        debit_val = txn.get('debit', '')
        if debit_val not in ('', None, 0, '0', '0.00'):
            try:
                debit_val = float(str(debit_val).replace(',', ''))
            except (ValueError, TypeError):
                pass
        else:
            debit_val = ''
        debit_cell = ws.cell(row=row, column=4, value=debit_val)
        debit_cell.border = thin_border
        debit_cell.number_format = '#,##0.00'

        credit_val = txn.get('credit', '')
        if credit_val not in ('', None, 0, '0', '0.00'):
            try:
                credit_val = float(str(credit_val).replace(',', ''))
            except (ValueError, TypeError):
                pass
        else:
            credit_val = ''
        credit_cell = ws.cell(row=row, column=5, value=credit_val)
        credit_cell.border = thin_border
        credit_cell.number_format = '#,##0.00'

        balance_val = txn.get('balance', '')
        if balance_val not in ('', None):
            try:
                balance_val = float(str(balance_val).replace(',', ''))
            except (ValueError, TypeError):
                pass
        balance_cell = ws.cell(row=row, column=6, value=balance_val)
        balance_cell.border = thin_border
        balance_cell.number_format = '#,##0.00'

        ws.cell(row=row, column=7, value=txn.get('ledger', '')).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20

    # Save
    base_name = os.path.splitext(filename)[0]
    output_name = f'{base_name}_parsed.xlsx'
    output_path = os.path.join(OUTPUT_FOLDER, output_name)
    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name=output_name)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
